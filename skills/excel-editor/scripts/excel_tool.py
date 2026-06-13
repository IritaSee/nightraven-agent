#!/home/mandabot/venv/bin/python3
import os
import sys
import argparse
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_excel_info(file_path):
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}
    
    try:
        xl = pd.ExcelFile(file_path)
        sheets = xl.sheet_names
        info = {"file": file_path, "sheets": []}
        for sheet in sheets:
            df = xl.parse(sheet, nrows=1)
            info["sheets"].append({
                "name": sheet,
                "columns": list(df.columns),
                "approx_rows": len(xl.parse(sheet))
            })
        return info
    except Exception as e:
        return {"error": str(e)}

def read_excel_sheet(file_path, sheet_name=None, limit=50, offset=0, output_format="markdown"):
    if not os.path.exists(file_path):
        print(f"Error: File not found {file_path}")
        return
    
    try:
        xl = pd.ExcelFile(file_path)
        if not sheet_name:
            sheet_name = xl.sheet_names[0]
        
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        total_rows = len(df)
        
        # Apply offset and limit
        df_subset = df.iloc[offset:offset+limit]
        
        if output_format == "json":
            print(df_subset.to_json(orient="records", indent=2))
        elif output_format == "csv":
            print(df_subset.to_csv(index=False))
        else: # markdown
            print(f"### Sheet: {sheet_name} (Showing rows {offset+1} to {min(offset+limit, total_rows)} of {total_rows})\n")
            print(df_subset.to_markdown(index=False))
    except Exception as e:
        print(f"Error reading sheet: {e}")

def write_excel_cell(file_path, sheet_name, cell, value):
    try:
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet if we are creating a specific one
            if sheet_name and sheet_name != "Sheet":
                wb.remove(wb.active)
        
        if not sheet_name:
            ws = wb.active
        else:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
        
        # Try to parse value as float, int, or json if possible
        parsed_value = value
        try:
            if value.lower() == 'true':
                parsed_value = True
            elif value.lower() == 'false':
                parsed_value = False
            elif value.startswith('[') or value.startswith('{'):
                parsed_value = json.loads(value)
            else:
                try:
                    parsed_value = int(value)
                except ValueError:
                    parsed_value = float(value)
        except Exception:
            pass
        
        ws[cell] = parsed_value
        wb.save(file_path)
        print(f"Successfully wrote '{parsed_value}' to cell {cell} in sheet '{ws.title}' of {file_path}")
    except Exception as e:
        print(f"Error writing cell: {e}")

def write_excel_row(file_path, sheet_name, row_data_json, row_index=None):
    try:
        row_data = json.loads(row_data_json)
        if not isinstance(row_data, list):
            print("Error: row-data must be a JSON list")
            return
        
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
            if sheet_name and sheet_name != "Sheet":
                wb.remove(wb.active)
                
        if not sheet_name:
            ws = wb.active
        else:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
        
        if row_index is not None:
            ws.insert_rows(row_index)
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_idx, value=val)
            print(f"Successfully inserted row at index {row_index} in sheet '{ws.title}'")
        else:
            ws.append(row_data)
            print(f"Successfully appended row to sheet '{ws.title}'")
            
        wb.save(file_path)
    except Exception as e:
        print(f"Error writing row: {e}")

def format_excel_range(file_path, sheet_name, cell_range, bold=None, font_color=None, bg_color=None, align=None):
    try:
        if not os.path.exists(file_path):
            print(f"Error: File not found {file_path}")
            return
            
        wb = openpyxl.load_workbook(file_path)
        if not sheet_name:
            ws = wb.active
        else:
            ws = wb[sheet_name]
            
        # Parse range
        cells = ws[cell_range]
        
        # If single cell, wrap in tuple to make it iterable
        if not isinstance(cells, tuple) and not isinstance(cells, list):
            cells = ((cells,),)
        elif len(cells) > 0 and not isinstance(cells[0], tuple) and not isinstance(cells[0], list):
            cells = (cells,)
            
        for row in cells:
            for cell in row:
                # Apply bold
                if bold is not None:
                    current_font = cell.font or Font()
                    cell.font = Font(name=current_font.name, size=current_font.size, 
                                     bold=bold, italic=current_font.italic, 
                                     color=current_font.color)
                
                # Apply font color (hex, e.g. FF0000)
                if font_color:
                    current_font = cell.font or Font()
                    cell.font = Font(name=current_font.name, size=current_font.size,
                                     bold=current_font.bold, italic=current_font.italic,
                                     color=font_color)
                                     
                # Apply background color (hex, e.g. FFFF00)
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    
                # Apply alignment
                if align:
                    cell.alignment = Alignment(horizontal=align)
                    
        wb.save(file_path)
        print(f"Successfully applied formatting to range {cell_range} in sheet '{ws.title}'")
    except Exception as e:
        print(f"Error formatting range: {e}")

def run_custom_code(file_path, code_str):
    try:
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
            
        # Prepare execution context
        loc = {"wb": wb, "openpyxl": openpyxl, "pd": pd}
        
        # Execute code
        exec(code_str, globals(), loc)
        
        wb.save(file_path)
        print(f"Successfully executed custom code on {file_path}")
    except Exception as e:
        print(f"Error executing custom code: {e}")

def main():
    parser = argparse.ArgumentParser(description="Excel Editor Helper Tool")
    parser.add_argument("--file", required=True, help="Path to Excel file")
    parser.add_argument("--sheet", help="Sheet name")
    parser.add_argument("--action", required=True, choices=["info", "read", "write-cell", "write-row", "format", "run-code"])
    
    # Read options
    parser.add_argument("--limit", type=int, default=50, help="Max rows to display")
    parser.add_argument("--offset", type=int, default=0, help="Row offset to start reading from")
    parser.add_argument("--format", default="markdown", choices=["markdown", "json", "csv"], help="Output format")
    
    # Write cell options
    parser.add_argument("--cell", help="Cell coordinate (e.g. A1) or range (e.g. A1:C3)")
    parser.add_argument("--value", help="Value to write")
    
    # Write row options
    parser.add_argument("--row-data", help="JSON array of row values")
    parser.add_argument("--row-index", type=int, help="Row index to insert at (1-based)")
    
    # Format options
    parser.add_argument("--bold", action="store_true", default=None, help="Set font to bold")
    parser.add_argument("--no-bold", action="store_false", dest="bold", help="Set font to not bold")
    parser.add_argument("--font-color", help="Font color in HEX (e.g. FF0000)")
    parser.add_argument("--bg-color", help="Background color in HEX (e.g. FFFF00)")
    parser.add_argument("--align", choices=["left", "center", "right"], help="Text alignment")
    
    # Run code options
    parser.add_argument("--code", help="Python code string to execute. 'wb' is preloaded.")
    
    args = parser.parse_args()
    
    if args.action == "info":
        info = get_excel_info(args.file)
        print(json.dumps(info, indent=2))
    elif args.action == "read":
        read_excel_sheet(args.file, args.sheet, args.limit, args.offset, args.format)
    elif args.action == "write-cell":
        if not args.cell or args.value is None:
            print("Error: --cell and --value are required for write-cell")
            sys.exit(1)
        write_excel_cell(args.file, args.sheet, args.cell, args.value)
    elif args.action == "write-row":
        if not args.row_data:
            print("Error: --row-data is required for write-row")
            sys.exit(1)
        write_excel_row(args.file, args.sheet, args.row_data, args.row_index)
    elif args.action == "format":
        if not args.cell:
            print("Error: --cell (or range) is required for format")
            sys.exit(1)
        format_excel_range(args.file, args.sheet, args.cell, args.bold, args.font_color, args.bg_color, args.align)
    elif args.action == "run-code":
        if not args.code:
            print("Error: --code is required for run-code")
            sys.exit(1)
        run_custom_code(args.file, args.code)

if __name__ == "__main__":
    main()
